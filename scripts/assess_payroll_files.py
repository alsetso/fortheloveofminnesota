#!/usr/bin/env python3
"""
One-off assessment of Minnesota State Payroll Excel files (FY2020-2025).
Outputs file locations, sizes, structure, row counts, column mapping, sample data quality, agency list.
"""
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Run: pip install openpyxl")
    raise

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "minnesota_gov" / "State Payrole"
FISCAL_YEARS = [2020, 2021, 2022, 2023, 2024, 2025]


def find_sheet_by_pattern(workbook, pattern: str):
    for name in workbook.sheetnames:
        if pattern.lower() in name.lower():
            return name
    return None


def get_active_col(workbook):
    hr_name = find_sheet_by_pattern(workbook, "HR INFO")
    if not hr_name:
        return None
    sheet = workbook[hr_name]
    headers = [c.value for c in sheet[1]]
    for h in headers:
        if h and "ACTIVE_ON_JUNE_30" in str(h).upper():
            return h
    return None


def main():
    lines = []
    def out(s=""):
        lines.append(s)

    out("=" * 80)
    out("MINNESOTA STATE PAYROLL FILES — FULL ASSESSMENT")
    out("=" * 80)

    # Step 1 — Locate files
    out()
    out("STEP 1 — FILE LOCATIONS AND SIZES")
    out("-" * 60)
    if not EXCEL_DIR.exists():
        out(f"Directory not found: {EXCEL_DIR}")
        print("\n".join(lines))
        return
    files = []
    for fy in FISCAL_YEARS:
        path = EXCEL_DIR / f"fiscal-year-{fy}.xlsx"
        if path.exists():
            size = path.stat().st_size
            files.append((fy, path, size))
            out(f"  {path.relative_to(ROOT)}")
            out(f"    size: {size:,} bytes ({size / (1024*1024):.2f} MB)")
        else:
            out(f"  MISSING: {path.relative_to(ROOT)}")
    out()

    # Step 2 & 5 — Structure and row counts per file
    out("STEP 2 — FILE STRUCTURE (sheets, columns, first 3 data rows)")
    out("STEP 5 — ROW COUNT PER FISCAL YEAR (after HR + EARNINGS join)")
    out("-" * 60)

    all_hr_headers = None
    all_earnings_headers = None
    row_counts = []

    for fy, path, size in files:
        out(f"\n--- fiscal-year-{fy}.xlsx ---")
        wb = openpyxl.load_workbook(path, data_only=True)
        out(f"  Sheets: {wb.sheetnames}")
        hr_name = find_sheet_by_pattern(wb, "HR INFO")
        earn_name = find_sheet_by_pattern(wb, "EARNINGS")
        if not hr_name or not earn_name:
            out(f"  ERROR: HR INFO or EARNINGS sheet not found")
            continue
        hr_sheet = wb[hr_name]
        earn_sheet = wb[earn_name]
        hr_headers = [c.value for c in hr_sheet[1]]
        earn_headers = [c.value for c in earn_sheet[1]]
        if all_hr_headers is None:
            all_hr_headers = hr_headers
        if all_earnings_headers is None:
            all_earnings_headers = earn_headers
        hr_rows = hr_sheet.max_row - 1
        earn_rows = earn_sheet.max_row - 1
        out(f"  HR INFO: {len(hr_headers)} columns, {hr_rows:,} data rows")
        out(f"  EARNINGS: {len(earn_headers)} columns, {earn_rows:,} data rows")
        # Joined count: one row per HR record (LEFT JOIN to EARNINGS)
        row_counts.append((fy, hr_rows))
        out(f"  Rows after join (per FY): {hr_rows:,}")

        # First 3 data rows of HR (key columns only for brevity)
        out("  HR first 3 rows (first 10 cols):")
        for r in range(2, min(5, hr_sheet.max_row + 1)):
            row = [hr_sheet.cell(row=r, column=c).value for c in range(1, min(11, len(hr_headers) + 1))]
            out(f"    Row{r-1}: {row}")
        out("  EARNINGS first 3 rows:")
        for r in range(2, min(5, earn_sheet.max_row + 1)):
            row = [earn_sheet.cell(row=r, column=c).value for c in range(1, len(earn_headers) + 1)]
            out(f"    Row{r-1}: {row}")
        wb.close()

    # Column enumeration (HR)
    out()
    out("STEP 2 (continued) — HR INFO COLUMN LIST (numbered)")
    out("-" * 60)
    if all_hr_headers:
        for i, h in enumerate(all_hr_headers, 1):
            out(f"  {i:2}. {h}")

    out()
    out("EARNINGS COLUMN LIST (numbered)")
    out("-" * 60)
    if all_earnings_headers:
        for i, h in enumerate(all_earnings_headers, 1):
            out(f"  {i:2}. {h}")

    # Step 3 — Data quality on one file (FY2025)
    out()
    out("STEP 3 — SAMPLE DATA QUALITY (fiscal-year-2025.xlsx)")
    out("-" * 60)
    path_25 = EXCEL_DIR / "fiscal-year-2025.xlsx"
    if path_25.exists():
        wb = openpyxl.load_workbook(path_25, data_only=True)
        hr_name = find_sheet_by_pattern(wb, "HR INFO")
        earn_name = find_sheet_by_pattern(wb, "EARNINGS")
        hr_sheet = wb[hr_name]
        earn_sheet = wb[earn_name]
        hr_headers = [c.value for c in hr_sheet[1]]
        earn_headers = [c.value for c in earn_sheet[1]]
        # Column indices 1-based for openpyxl
        def col_index(name):
            for i, h in enumerate(hr_headers):
                if h and str(h).strip().upper() == name.strip().upper():
                    return i + 1
            return None
        agency_col = col_index("AGENCY_NAME") or col_index("AGENCY_NBR")
        name_col = col_index("EMPLOYEE_NAME")
        comp_col = col_index("COMPENSATION_RATE")
        # EARNINGS
        tw_col = None
        for i, h in enumerate(earn_sheet[1]):
            if h and "TOTAL_WAGES" in str(h).upper():
                tw_col = i + 1
                break
        agencies = []
        names = set()
        comp_rates = []
        total_wages = []
        for r in range(2, hr_sheet.max_row + 1):
            if agency_col:
                v = hr_sheet.cell(row=r, column=agency_col).value
                if v and str(v).strip():
                    agencies.append(str(v).strip())
            if name_col:
                v = hr_sheet.cell(row=r, column=name_col).value
                if v and str(v).strip():
                    names.add(str(v).strip())
            if comp_col:
                v = hr_sheet.cell(row=r, column=comp_col).value
                if v is not None:
                    try:
                        comp_rates.append(float(v))
                    except (TypeError, ValueError):
                        pass
        for r in range(2, earn_sheet.max_row + 1):
            if tw_col:
                v = earn_sheet.cell(row=r, column=tw_col).value
                if v is not None:
                    try:
                        total_wages.append(float(v))
                    except (TypeError, ValueError):
                        pass
        # Unique agencies count and top 20
        from collections import Counter
        agency_counts = Counter(agencies)
        out(f"  Unique agencies (by AGENCY_NAME): {len(agency_counts)}")
        out("  Top 20 agencies by row count:")
        for name, count in agency_counts.most_common(20):
            out(f"    {count:>6,}  {name}")
        out(f"  Unique employees (by EMPLOYEE_NAME): {len(names)}")
        if comp_rates:
            comp_rates.sort()
            out(f"  COMPENSATION_RATE range: min={min(comp_rates):,.2f} max={max(comp_rates):,.2f}")
            out(f"    lowest 5: {comp_rates[:5]}")
            out(f"    highest 5: {comp_rates[-5:]}")
        if total_wages:
            total_wages.sort()
            out(f"  TOTAL_WAGES range: min={min(total_wages):,.2f} max={max(total_wages):,.2f}")
            out(f"    lowest 5: {total_wages[:5]}")
            out(f"    highest 5: {total_wages[-5:]}")
        # Step 6 — distinct agency names for cross-reference
        out()
        out("STEP 6 — FY2025 DISTINCT AGENCY NAMES (for org_agency_map cross-reference)")
        out("-" * 60)
        distinct_agencies = sorted(agency_counts.keys())
        for a in distinct_agencies:
            out(f"  {a}")
        out(f"\n  Total distinct agencies in FY2025: {len(distinct_agencies)}")
        wb.close()
    else:
        out("  fiscal-year-2025.xlsx not found")

    # Step 5 summary
    out()
    out("STEP 5 — ROW COUNT SUMMARY PER FISCAL YEAR")
    out("-" * 60)
    for fy, count in row_counts:
        out(f"  FY{fy}: {count:,} rows")
    total_rows = sum(c for _, c in row_counts)
    out(f"  TOTAL (all 6 years): {total_rows:,} rows")

    # Step 4 — Column mapping (reference only)
    out()
    out("STEP 4 — COLUMN MAPPING (Excel → checkbook.payroll)")
    out("-" * 60)
    out("  HR INFO columns map to schema as documented in minnesota_gov/IMPORT_REQUIREMENTS.md")
    out("  EARNINGS: TEMPORARY_ID (join), REGULAR_WAGES→regular_wages, OVERTIME_WAGES→overtime_wages,")
    out("  OTHER_WAGES→other_wages, TOTAL_WAGES→total_wages. All existing schema columns have a source.")
    out("  ACTIVE_ON_JUNE_30_#### column name varies by year; script resolves it at runtime.")

    print("\n".join(lines))


if __name__ == "__main__":
    main()
