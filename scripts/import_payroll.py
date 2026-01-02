#!/usr/bin/env python3
"""
Import payroll data from Excel files into checkbook.payroll table.

Usage:
    python scripts/import_payroll.py

Environment variables required:
    NEXT_PUBLIC_SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY
"""

import os
import sys
from pathlib import Path
from typing import Optional, Dict, List
from dotenv import load_dotenv

try:
    from supabase import create_client, Client
    import openpyxl
except ImportError:
    print("Error: Required packages not installed.")
    print("Run: pip install supabase python-dotenv openpyxl")
    sys.exit(1)

# Load environment variables
env_path = Path(__file__).parent.parent / '.env.local'
if env_path.exists():
    load_dotenv(env_path)
else:
    load_dotenv()

# Configuration
BATCH_SIZE = 1000
EXCEL_DIR = Path(__file__).parent.parent / "minnesota_gov" / "State Payrole"
FISCAL_YEARS = [2020]  # Testing with 2020 first


def parse_integer(value) -> Optional[int]:
    """Parse value to integer, return None if invalid."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value or value == '-':
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None
    return None


def parse_float(value) -> Optional[float]:
    """Parse value to float, return None if invalid."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value or value == '-':
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    return None


def parse_text(value) -> Optional[str]:
    """Normalize text field: empty strings become None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value).strip() if value else None
    if isinstance(value, str):
        value = value.strip()
        return value if value and value != '-' else None
    return None


def parse_last_hire_date(value) -> Optional[str]:
    """Parse LAST_HIRE_DATE which can be integer or '-' string."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(value))
    if isinstance(value, str):
        value = value.strip()
        if value == '-' or not value:
            return None
        try:
            # Try to parse as integer
            return str(int(float(value)))
        except (ValueError, TypeError):
            return value if value else None
    return None


def find_sheet_by_pattern(workbook, pattern: str) -> Optional[str]:
    """Find sheet name containing pattern."""
    for sheet_name in workbook.sheetnames:
        if pattern.lower() in sheet_name.lower():
            return sheet_name
    return None


def get_active_column_name(workbook, fiscal_year: int) -> Optional[str]:
    """Get the ACTIVE_ON_JUNE_30 column name (varies by year)."""
    hr_sheet_name = find_sheet_by_pattern(workbook, 'HR INFO')
    if not hr_sheet_name:
        return None
    
    sheet = workbook[hr_sheet_name]
    # Get header row
    header_row = [cell.value for cell in sheet[1]]
    for col_name in header_row:
        if col_name and 'ACTIVE_ON_JUNE_30' in str(col_name).upper():
            return col_name
    return None


def parse_hr_row(row: List, headers: List[str], active_col_name: str) -> Optional[Dict]:
    """Parse HR INFO row into dictionary."""
    try:
        data = {}
        for i, header in enumerate(headers):
            if i < len(row):
                data[header] = row[i]
            else:
                data[header] = None
        
        temporary_id = parse_text(data.get('TEMPORARY_ID'))
        if not temporary_id:
            return None
        
        return {
            'temporary_id': temporary_id,
            'record_nbr': parse_integer(data.get('RECORD_NBR')),
            'employee_name': parse_text(data.get('EMPLOYEE_NAME')),
            'agency_nbr': parse_text(data.get('AGENCY_NBR')),
            'agency_name': parse_text(data.get('AGENCY_NAME')),
            'department_nbr': parse_text(data.get('DEPARTMENT_NBR')),
            'department_name': parse_text(data.get('DEPARTMENT_NAME')),
            'branch_code': parse_text(data.get('BRANCH_CODE')),
            'branch_name': parse_text(data.get('BRANCH_NAME')),
            'job_code': parse_text(data.get('JOB_CODE')),
            'job_title': parse_text(data.get('JOB_TITLE')),
            'location_nbr': parse_text(data.get('LOCATION_NBR')),
            'location_name': parse_text(data.get('LOCATION_NAME')),
            'location_county_name': parse_text(data.get('LOCATION_COUNTY_NAME')),
            'reg_temp_code': parse_text(data.get('REG_TEMP_CODE')),
            'reg_temp_desc': parse_text(data.get('REG_TEMP_DESC')),
            'classified_code': parse_text(data.get('CLASSIFIED_CODE')),
            'classified_desc': parse_text(data.get('CLASSIFIED_DESC')),
            'original_hire_date': parse_integer(data.get('ORIGINAL_HIRE_DATE')),
            'last_hire_date': parse_last_hire_date(data.get('LAST_HIRE_DATE')),
            'job_entry_date': parse_integer(data.get('JOB_ENTRY_DATE')),
            'full_part_time_code': parse_text(data.get('FULL_PART_TIME_CODE')),
            'full_part_time_desc': parse_text(data.get('FULL_PART_TIME_DESC')),
            'salary_plan_grid': parse_text(data.get('SALARY_PLAN_GRID')),
            'salary_grade_range': parse_integer(data.get('SALARY_GRADE_RANGE')),
            'max_salary_step': parse_integer(data.get('MAX_SALARY_STEP')),
            'compensation_rate': parse_float(data.get('COMPENSATION_RATE')),
            'comp_frequency_code': parse_text(data.get('COMP_FREQUENCY_CODE')),
            'comp_frequency_desc': parse_text(data.get('COMP_FREQUENCY_DESC')),
            'position_fte': parse_float(data.get('POSITION_FTE')),
            'bargaining_unit_nbr': parse_integer(data.get('BARGAINING_UNIT_NBR')),
            'bargaining_unit_name': parse_text(data.get('BARGAINING_UNIT_NAME')),
            'active_on_june_30': parse_text(data.get(active_col_name)) if active_col_name else None,
        }
    except Exception as e:
        print(f"    Error parsing HR row: {e}")
        return None


def parse_earnings_row(row: List, headers: List[str]) -> Optional[Dict]:
    """Parse EARNINGS row into dictionary."""
    try:
        data = {}
        for i, header in enumerate(headers):
            if i < len(row):
                data[header] = row[i]
            else:
                data[header] = None
        
        temporary_id = parse_text(data.get('TEMPORARY_ID'))
        if not temporary_id:
            return None
        
        return {
            'temporary_id': temporary_id,
            'regular_wages': parse_float(data.get('REGULAR_WAGES')) or 0.0,
            'overtime_wages': parse_float(data.get('OVERTIME_WAGES')) or 0.0,
            'other_wages': parse_float(data.get('OTHER_WAGES')) or 0.0,
            'total_wages': parse_float(data.get('TOTAL_WAGES')) or 0.0,
        }
    except Exception as e:
        print(f"    Error parsing earnings row: {e}")
        return None


def import_payroll_file(supabase: Client, file_path: Path, fiscal_year: int) -> tuple[int, int]:
    """Import a single payroll Excel file."""
    print(f"  Processing {file_path.name}...")
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Find HR INFO and EARNINGS sheets
        hr_sheet_name = find_sheet_by_pattern(workbook, 'HR INFO')
        earnings_sheet_name = find_sheet_by_pattern(workbook, 'EARNINGS')
        
        if not hr_sheet_name:
            print(f"    Error: HR INFO sheet not found")
            return 0, 0
        if not earnings_sheet_name:
            print(f"    Error: EARNINGS sheet not found")
            return 0, 0
        
        # Get active column name
        active_col_name = get_active_column_name(workbook, fiscal_year)
        
        # Read HR INFO sheet
        hr_sheet = workbook[hr_sheet_name]
        hr_headers = [cell.value for cell in hr_sheet[1]]
        hr_data = {}
        
        print(f"    Reading HR INFO sheet ({hr_sheet.max_row - 1} rows)...")
        for row in hr_sheet.iter_rows(min_row=2, values_only=True):
            parsed = parse_hr_row(list(row), hr_headers, active_col_name)
            if parsed and parsed['temporary_id']:
                hr_data[parsed['temporary_id']] = parsed
        
        # Read EARNINGS sheet
        earnings_sheet = workbook[earnings_sheet_name]
        earnings_headers = [cell.value for cell in earnings_sheet[1]]
        earnings_data = {}
        
        print(f"    Reading EARNINGS sheet ({earnings_sheet.max_row - 1} rows)...")
        for row in earnings_sheet.iter_rows(min_row=2, values_only=True):
            parsed = parse_earnings_row(list(row), earnings_headers)
            if parsed and parsed['temporary_id']:
                earnings_data[parsed['temporary_id']] = parsed
        
        # Join HR and EARNINGS data
        print(f"    Joining HR and EARNINGS data...")
        records = []
        for temp_id, hr_record in hr_data.items():
            earnings_record = earnings_data.get(temp_id, {})
            combined = {**hr_record}
            combined.update({
                'fiscal_year': str(fiscal_year),  # Add fiscal year as text to each record
                'regular_wages': earnings_record.get('regular_wages', 0.0),
                'overtime_wages': earnings_record.get('overtime_wages', 0.0),
                'other_wages': earnings_record.get('other_wages', 0.0),
                'total_wages': earnings_record.get('total_wages', 0.0),
            })
            records.append(combined)
        
        if not records:
            print(f"  No valid records found in {file_path.name}")
            return 0, 0
        
        # Batch insert
        total_inserted = 0
        skipped = 0
        for i in range(0, len(records), BATCH_SIZE):
            batch = records[i:i + BATCH_SIZE]
            try:
                # Insert batch (temporary_id is not unique across years, so regular insert)
                result = supabase.table('payroll').insert(batch).execute()
                inserted = len(result.data) if result.data else len(batch)
                total_inserted += inserted
                print(f"    Inserted batch: {inserted} records (total: {total_inserted}/{len(records)})")
            except Exception as e:
                print(f"    Error inserting batch {i//BATCH_SIZE + 1}: {e}")
                skipped += len(batch)
        
        return total_inserted, skipped
    
    except FileNotFoundError:
        print(f"  Error: File not found: {file_path}")
        return 0, 0
    except Exception as e:
        print(f"  Error processing {file_path.name}: {e}")
        import traceback
        traceback.print_exc()
        return 0, 0


def main():
    """Main import function."""
    # Validate environment
    supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
    
    if not supabase_url:
        print("Error: NEXT_PUBLIC_SUPABASE_URL not set in environment")
        sys.exit(1)
    
    if not supabase_key:
        print("Error: SUPABASE_SERVICE_ROLE_KEY not set in environment")
        sys.exit(1)
    
    # Initialize Supabase client
    try:
        supabase: Client = create_client(supabase_url, supabase_key)
    except Exception as e:
        print(f"Error creating Supabase client: {e}")
        sys.exit(1)
    
    # Validate Excel directory
    if not EXCEL_DIR.exists():
        print(f"Error: Excel directory not found: {EXCEL_DIR}")
        sys.exit(1)
    
    print("=" * 60)
    print("Payroll Import Script")
    print("=" * 60)
    print(f"Excel Directory: {EXCEL_DIR}")
    print(f"Batch Size: {BATCH_SIZE}")
    print()
    
    # Process each fiscal year
    total_inserted = 0
    total_skipped = 0
    
    for fiscal_year in FISCAL_YEARS:
        file_path = EXCEL_DIR / f"fiscal-year-{fiscal_year}.xlsx"
        
        if not file_path.exists():
            print(f"⚠️  Skipping FY{fiscal_year}: File not found")
            continue
        
        print(f"📁 Fiscal Year {fiscal_year}:")
        inserted, skipped = import_payroll_file(supabase, file_path, fiscal_year)
        total_inserted += inserted
        total_skipped += skipped
        print()
    
    # Summary
    print("=" * 60)
    print("Import Summary")
    print("=" * 60)
    print(f"Total records inserted/updated: {total_inserted:,}")
    print(f"Total rows skipped: {total_skipped:,}")
    print("=" * 60)


if __name__ == '__main__':
    main()

