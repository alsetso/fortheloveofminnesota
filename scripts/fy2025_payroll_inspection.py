#!/usr/bin/env python3
"""
FY2025 payroll file inspection only. No DB writes.
Output: file details, headers, first 3 rows, null counts, distinct agencies, uniqueness of (temporary_id, record_nbr).
"""
from pathlib import Path
import json

try:
    import openpyxl
except ImportError:
    print("Run: pip install openpyxl")
    raise

ROOT = Path(__file__).resolve().parent.parent
FILE = ROOT / "minnesota_gov" / "State Payrole" / "fiscal-year-2025.xlsx"


def find_sheet(workbook, pattern):
    for name in workbook.sheetnames:
        if pattern.lower() in name.lower():
            return name
    return None


def main():
    if not FILE.exists():
        print(f"File not found: {FILE}")
        return

    # File details
    size = FILE.stat().st_size
    print("=== STEP 1 — FILE INSPECTION ===")
    print(f"Exact file path: {FILE.resolve()}")
    print(f"File name: {FILE.name}")
    print(f"File size: {size:,} bytes ({size / (1024*1024):.2f} MB)")
    print("Note: This is an Excel (.xlsx) file. wc -l / head / cut do not apply; row counts below are from opening the workbook.")

    wb = openpyxl.load_workbook(FILE, data_only=True)
    hr_name = find_sheet(wb, "HR INFO")
    earn_name = find_sheet(wb, "EARNINGS")
    if not hr_name or not earn_name:
        print("HR INFO or EARNINGS sheet not found")
        return

    hr = wb[hr_name]
    earn = wb[earn_name]
    hr_headers = [c.value for c in hr[1]]
    earn_headers = [c.value for c in earn[1]]
    hr_data_rows = hr.max_row - 1
    earn_data_rows = earn.max_row - 1
    # After join: one row per HR row (LEFT JOIN EARNINGS)
    total_rows_after_join = hr_data_rows
    print(f"Total data rows (HR INFO): {hr_data_rows:,} (header row excluded)")
    print(f"Total data rows (EARNINGS): {earn_data_rows:,}")
    print(f"Total row count including header (conceptual): 1 header + {total_rows_after_join:,} = {1 + total_rows_after_join:,} (one row per HR record after join)")

    print("\n--- HR INFO: every column name exactly as in header ---")
    for i, h in enumerate(hr_headers, 1):
        print(f"  {i:2}. {repr(h)}")

    print("\n--- EARNINGS: every column name exactly as in header ---")
    for i, h in enumerate(earn_headers, 1):
        print(f"  {i:2}. {repr(h)}")

    print("\n--- First 3 data rows (HR INFO), all columns ---")
    for r in range(2, min(5, hr.max_row + 1)):
        row = [hr.cell(row=r, column=c).value for c in range(1, len(hr_headers) + 1)]
        print(f"  Row {r-1}: {row}")

    print("\n--- First 3 data rows (EARNINGS), all columns ---")
    for r in range(2, min(5, earn.max_row + 1)):
        row = [earn.cell(row=r, column=c).value for c in range(1, len(earn_headers) + 1)]
        print(f"  Row {r-1}: {row}")

    # Column indices (1-based for openpyxl)
    def col_idx(sheet, name):
        headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        for i, h in enumerate(headers):
            if h and str(h).strip().upper() == name.strip().upper():
                return i + 1
        return None

    emp_col = col_idx(hr, "EMPLOYEE_NAME")
    agency_col = col_idx(hr, "AGENCY_NAME")
    temp_id_col = col_idx(hr, "TEMPORARY_ID")
    rec_nbr_col = col_idx(hr, "RECORD_NBR")
    tw_col = col_idx(earn, "TOTAL_WAGES")

    # Build earnings by temporary_id (one earnings row per temp_id in source; we join to HR so total_wages per HR row = from earnings or 0)
    earnings_by_id = {}
    for r in range(2, earn.max_row + 1):
        tid = earn.cell(row=r, column=1).value
        if tid is not None:
            tid = str(tid).strip()
        tw = earn.cell(row=r, column=5).value if earn_headers and len(earn_headers) >= 5 else None
        if tid:
            earnings_by_id[tid] = tw  # last occurrence if duplicate temp_id in EARNINGS

    # Step 3 — Null counts (on the joined conceptual row: employee_name, agency_name from HR; total_wages from EARNINGS for that temp_id, or null/empty)
    null_employee = 0
    null_agency = 0
    null_total_wages = 0  # count HR rows where we'd have no earnings or earnings TOTAL_WAGES is null/empty
    agencies = []
    keys = []  # (temporary_id, record_nbr)
    for r in range(2, hr.max_row + 1):
        emp = hr.cell(row=r, column=emp_col).value if emp_col else None
        ag = hr.cell(row=r, column=agency_col).value if agency_col else None
        tid = hr.cell(row=r, column=temp_id_col).value if temp_id_col else None
        rn = hr.cell(row=r, column=rec_nbr_col).value if rec_nbr_col else None
        if emp is None or (isinstance(emp, str) and not emp.strip()):
            null_employee += 1
        if ag is None or (isinstance(ag, str) and not ag.strip()):
            null_agency += 1
        if tid is not None:
            tid = str(tid).strip()
        tw_val = earnings_by_id.get(tid) if tid else None
        if tw_val is None or (isinstance(tw_val, str) and (not tw_val.strip() or tw_val.strip() == '-')):
            null_total_wages += 1
        if ag is not None and str(ag).strip():
            agencies.append(str(ag).strip())
        if tid is not None and rn is not None:
            keys.append((str(tid).strip(), rn if isinstance(rn, (int, float)) else str(rn)))

    print("\n=== STEP 3 — DATA QUALITY (FY2025, joined view) ===")
    print(f"Null or empty employee_name: {null_employee:,}")
    print(f"Null or empty agency_name: {null_agency:,}")
    print(f"Null/empty/missing total_wages (no EARNINGS or TOTAL_WAGES blank/dash): {null_total_wages:,}")

    distinct_agencies = sorted(set(agencies))
    print(f"\nDistinct agency names in file: {len(distinct_agencies)}")
    print("Full list of distinct agency names:")
    for a in distinct_agencies:
        print(f"  {a}")

    # Step 5 — Uniqueness of (temporary_id, record_nbr)
    print("\n=== STEP 5 — UPSERT KEY ===")
    print("Source has TEMPORARY_ID (col 1) and RECORD_NBR (col 2). Same person can have multiple rows (multiple RECORD_NBR per TEMPORARY_ID).")
    unique_keys = set(keys)
    print(f"Rows in HR: {hr_data_rows:,}")
    print(f"Unique (temporary_id, record_nbr) pairs: {len(unique_keys):,}")
    if len(keys) != len(unique_keys):
        print("WARNING: (temporary_id, record_nbr) is NOT unique — duplicate pairs exist.")
    else:
        print("(temporary_id, record_nbr) is UNIQUE across all HR rows.")
    print("For upsert: (temporary_id, record_nbr, fiscal_year) is the natural composite key per row.")

    # Output agency list as SQL values for Step 4
    print("\n=== AGENCY LIST FOR STEP 4 (SQL values) ===")
    # Escape single quotes for SQL
    def esc(s):
        return s.replace("'", "''")
    values = ", ".join(f"('{esc(a)}')" for a in distinct_agencies)
    print(f"(-- Paste into: ... from (values {values[:80]}...) as src(agency_name) ...)")
    # Write full list to a file for easy paste
    out_path = ROOT / "scripts" / "fy2025_agency_list.txt"
    with open(out_path, "w") as f:
        f.write("\n".join(distinct_agencies))
    print(f"Full list also written to {out_path}")

    wb.close()


if __name__ == "__main__":
    main()
